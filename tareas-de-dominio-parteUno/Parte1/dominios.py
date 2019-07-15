from openpyxl import load_workbook
from get_domains import get_domains
import os


def path_absolute():
    os.chdir(r'dominios_existentes')
    cwd = os.getcwd()
    return cwd


def get_domains_list():
    """
    Obtiene los dominios que estan para trabajar y los devuelve en una lista
    """
    os.chdir(r'dominios_existentes')
    cwd = os.getcwd()
    file_path = r"/Dominios.xlsx"
    archivo = cwd + file_path
    filesheet = load_workbook(archivo, read_only=True)
    sheet = filesheet.active
    filas = sheet.max_row
    dominios = get_domains(filas, sheet)  #get_domains(filas, sheet)
    return dominios


def path_screenshots():
    try:
        os.chdir('../screenshots')
        saved_path = os.getcwd()
        return saved_path
    except FileNotFoundError as e:
        logging.error(e, "¡ERROR! Directorio no encontrado  :( ")
