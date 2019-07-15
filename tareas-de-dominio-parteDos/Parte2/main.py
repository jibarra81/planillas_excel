from get_info import create_sheet
from get_info import get_file, get_is_current, get_is_proposed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import logging


wb = Workbook()
filesheet = "planillas_generadas/resultado.xlsx"
create_sheet(wb, filesheet)
archivo = get_file()[0]
planilla_activa = wb.active


def encabezado():
    """ Crea el encabezado de la planilla """
    try:
        planilla_activa['A1'].font = Font(bold=True, size=11)
        planilla_activa['B1'].font = Font(bold=True, size=11)
        planilla_activa['C1'].font = Font(bold=True, size=11)
        planilla_activa['D1'].font = Font(bold=True, size=11)

        planilla_activa['A1'].alignment = Alignment(horizontal='center')
        planilla_activa['B1'].alignment = Alignment(horizontal='center')
        planilla_activa['C1'].alignment = Alignment(horizontal='center')
        planilla_activa['D1'].alignment = Alignment(horizontal='center')

        planilla_activa['A1'] = "Dominios"
        planilla_activa['B1'] = "Current"
        planilla_activa['C1'] = "Proposed"
        planilla_activa['D1'] = "Total"
    except Exception as e:
        logging.error(e, '¡ERROR! - No se pudo crear el encabezado')


def generar_planilla():
    try:
        nombre_archivo = get_file()[1]
        logging.info("Se obtiene el nombre del archivo")
        for i in range(len(archivo)):
            sheetML = load_workbook(archivo[i])
            pestania = sheetML.active.title
            sh = sheetML[pestania]
            filas = sh.max_row
            titulo = str(nombre_archivo[i][10:-37])
            current = get_is_current(filas, sh)
            proposed = get_is_proposed(filas, sh)
            datos = [(titulo.strip(), int(current), int(proposed), int(current) + int(proposed))]
            encabezado()
            for registro in datos:
                logging.info(registro)
                planilla_activa.append(registro)
        logging.info('Se genero la planilla')
    except Exception as e:
        logging.error(e, '¡ERROR! - No se pudo generar la planilla')


logging.debug("Comienza la ejecucion del programa")
generar_planilla()
wb.save(filesheet)
logging.info("Finaliza la ejecucion")
logging.info("--------------------------------------------------------")
