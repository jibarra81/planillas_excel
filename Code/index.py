from get_info import create_sheet
from get_info import get_file, get_is_current, get_is_proposed
from openpyxl import Workbook, load_workbook


wb = Workbook()
filesheet = "../Planillas_generadas/planilla.xlsx"

create_sheet(wb, filesheet)
archivo = get_file()[0]
nombre_archivo = get_file()[1]
sheetML = load_workbook(archivo[0])
pestania = sheetML.active.title
sh = sheetML[pestania]
filas = sh.max_row
planilla_activa = wb.active
titulo = str(nombre_archivo[0])

current = get_is_current(filas, sh)
proposed = get_is_proposed(filas, sh)

datos = [('Dominio:', titulo),
         ("Planillas Current: ", str(current)),
         ("Planillas Proposed: ", str(proposed)),
         ("Total: ", int(current)+int(proposed)),
         ("##########", "##########")]

for registro in datos:
    print(registro)
    planilla_activa.append(registro)

wb.save(filesheet)
