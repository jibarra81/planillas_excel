from get_info import create_sheet
from get_info import get_file, get_is_current, get_is_proposed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


wb = Workbook()
filesheet = "../Planillas_generadas/planilla.xlsx"
create_sheet(wb, filesheet)
archivo = get_file()[0]
planilla_activa = wb.active


def encabezado():
    planilla_activa['A1'].font = Font(bold=True, size=11)
    planilla_activa['B1'].font = Font(bold=True, size=11)
    planilla_activa['C1'].font = Font(bold=True, size=11)
    planilla_activa['D1'].font = Font(bold=True, size=11)

    planilla_activa['A1'].alignment = Alignment(horizontal='center')
    planilla_activa['B1'].alignment = Alignment(horizontal='center')
    planilla_activa['C1'].alignment = Alignment(horizontal='center')
    planilla_activa['D1'].alignment = Alignment(horizontal='center')

    planilla_activa['A1'] = "Dominios"
    planilla_activa['B1'] = "Current"
    planilla_activa['C1'] = "Proposed"
    planilla_activa['D1'] = "Total"


def generar_planilla():
    nombre_archivo = get_file()[1]
    for i in range(len(archivo)):
        sheetML = load_workbook(archivo[i])
        pestania = sheetML.active.title
        sh = sheetML[pestania]
        filas = sh.max_row
        titulo = str(nombre_archivo[i][10:-37])
        current = get_is_current(filas, sh)
        proposed = get_is_proposed(filas, sh)
        datos = [(titulo.strip(), int(current), int(proposed), int(current)+int(proposed))]
        encabezado()
        for registro in datos:
            print(registro)
            planilla_activa.append(registro)


generar_planilla()
wb.save(filesheet)
